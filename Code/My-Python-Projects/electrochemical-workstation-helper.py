import xlsxwriter as xw
import os
import time
from openpyxl import load_workbook
from tqdm import tqdm


# 1.用户自定义存储电池有效面积变量
while True:
    device_area = input('请输入电池有效面积：')
    try:
        if type(eval(device_area)) == float or type(eval(device_area)) == int:
            confirm_flag = input('按1确认电池有效面积，按任意键重新输入...：')
            if confirm_flag == '1':
                print(f'确认完毕，你的电池有效面积是 {device_area}')
                break
    except Exception as error_message:
        print('你输入的不是有效数据类型！')

device_area = eval(device_area)

time_start = time.process_time()

# 2.规划汇总数据文件名
folder = str(os.getcwd())
index = folder.rfind('\\')
result_name = folder[index+1:] + '.xlsx'


# 3.创建汇总数据工作簿
result_workbook = xw.Workbook(result_name)
format = result_workbook.add_format({'align': 'center', 'font_color': 'red', 'bold': True})
format1 = result_workbook.add_format({'align': 'center'})
format2 = result_workbook.add_format({'align': 'center', 'font_color': 'blue'})
result_worksheet = result_workbook.add_worksheet('result')
result_worksheet.set_column(0, 6, 15)
result_worksheet.set_column(7, 7, 25)
result_worksheet.write(0, 0, 'file name', format)
result_worksheet.write(0, 1, 'Mode', format)
result_worksheet.write(0, 2, 'Voc (V)', format)
result_worksheet.write(0, 3, 'Jsc (mA/cm^2)', format)
result_worksheet.write(0, 4, 'FF (%)', format)
result_worksheet.write(0, 5, 'PCE (%)', format)
result_worksheet.write(0, 6, 'Device area (cm^2)', format)
result_worksheet.write(0, 6, 'Device area (cm^2)', format)
result_worksheet.write(0, 7, 'Author: umaru', format)
result_worksheet.write(1, 7, 'From', format)
result_worksheet.write(2, 7, '1 kWh Group', format)
result_worksheet.write(3, 7, 'Funsom - Soochow.U', format)


# 4.读取并筛选数据
files_list = os.listdir()
find_Voc = []
find_Jsc = []
find_P_max = []
single_data = []
data = []
for i in files_list:
    if i == 'electrochemical-workstation-helper.py' or i == result_name or i == 'electrochemical-workstation-helper.exe':
        pass
    else:
        single_data.append(i)
        i_workbook = load_workbook(i)
        i_sheet = i_workbook.worksheets[0]
        row_num = i_sheet.max_row
        cell = i_sheet.cell(2, 1)
        if cell.value < 0:
            single_data.append('Forward')
            for row in range(2, row_num+1):
                if_V = i_sheet.cell(row, 4).value
                if_I = i_sheet.cell(row, 3).value
                if if_I < 0 and if_V > 0:
                    if_P_max = (abs(if_I) * 1000 / device_area) * if_V
                    find_P_max.append(if_P_max)
                if find_Jsc == []:
                    if i_sheet.cell(row, 4).value > 0:
                        find_Jsc.append([i_sheet.cell(row, 4).value, i_sheet.cell(row, 3).value])
                        find_Jsc.append([i_sheet.cell(row-1, 4).value, i_sheet.cell(row-1, 3).value])
                        correct_Jsc = find_Jsc[0][1] - find_Jsc[0][0] * ((find_Jsc[1][1] - find_Jsc[0][1]) / (find_Jsc[1][0] - find_Jsc[0][0]))
                        correct_Jsc = round(abs(correct_Jsc * 1000 / device_area), 6)
                        single_data.append(correct_Jsc)
                    else:
                        pass
                elif i_sheet.cell(row, 3).value > 0:
                    find_Voc.append([i_sheet.cell(row, 4).value, i_sheet.cell(row, 3).value])
                    find_Voc.append([i_sheet.cell(row - 1, 4).value, i_sheet.cell(row - 1, 3).value])
                    correct_Voc = find_Voc[1][0] - find_Voc[1][1] * ((find_Voc[0][0] - find_Voc[1][0]) / (find_Voc[0][1] - find_Voc[1][1]))
                    correct_Voc = round(abs(correct_Voc), 6)
                    single_data.append(correct_Voc)
                    single_data.append(max(find_P_max))
                    data.append(single_data)
                    i_workbook.save(i)
                    single_data = []
                    find_Voc = []
                    find_Jsc = []
                    find_P_max =[]
                    break

        if cell.value > 0:
            single_data.append('Reverse')
            for row in range(2, row_num + 1):
                if_V = i_sheet.cell(row, 4).value
                if_I = i_sheet.cell(row, 3).value
                if if_I < 0 and if_V > 0:
                    if_P_max = (abs(if_I) * 1000 / device_area) * if_V
                    find_P_max.append(if_P_max)
                if find_Voc == []:
                    if i_sheet.cell(row, 3).value < 0:
                        find_Voc.append([i_sheet.cell(row, 4).value, i_sheet.cell(row, 3).value])
                        find_Voc.append([i_sheet.cell(row - 1, 4).value, i_sheet.cell(row - 1, 3).value])
                        correct_Voc = find_Voc[1][0] - find_Voc[1][1] * ((find_Voc[0][0] - find_Voc[1][0]) / (find_Voc[0][1] - find_Voc[1][1]))
                        correct_Voc = round(abs(correct_Voc), 6)
                        single_data.append(correct_Voc)
                    else:
                        pass
                elif i_sheet.cell(row, 4).value < 0:
                    find_Jsc.append([i_sheet.cell(row, 4).value, i_sheet.cell(row, 3).value])
                    find_Jsc.append([i_sheet.cell(row - 1, 4).value, i_sheet.cell(row - 1, 3).value])
                    correct_Jsc = find_Jsc[0][1] - find_Jsc[0][0] * ((find_Jsc[1][1] - find_Jsc[0][1]) / (find_Jsc[1][0] - find_Jsc[0][0]))
                    correct_Jsc = round(abs(correct_Jsc * 1000 / device_area), 6)
                    single_data.append(correct_Jsc)
                    single_data.append(max(find_P_max))
                    data.append(single_data)
                    i_workbook.save(i)
                    single_data = []
                    find_Voc = []
                    find_Jsc = []
                    find_P_max = []
                    break


# 5.写入数据
n =1
for m in data:
    if n <= len(data):
            # 写入文件名
            result_worksheet.write(n, 0, m[0], format1)
            if m[1] == 'Forward':
                # 写入电压
                result_worksheet.write(n, 2, m[3], format1)
                # 写入电流
                result_worksheet.write(n, 3, m[2], format1)
                # 写入扫描模式 -- Forward
                result_worksheet.write(n, 1, m[1], format1)
            else:
                # 写入电压
                result_worksheet.write(n, 2, m[2], format1)
                # 写入电流
                result_worksheet.write(n, 3, m[3], format1)
                # 写入扫描模式 -- Reverse -- 蓝色标记
                result_worksheet.write(n, 1, m[1], format2)
            # 重新计算 FF
            data_FF = round((m[4] / (m[2] * m[3])), 6)
            result_worksheet.write(n, 4, data_FF, format1)
            # 重新计算 PCE
            data_PCE = round((m[2] * m[3] * data_FF), 6)
            result_worksheet.write(n, 5, data_PCE, format1)
            # 写入有效面积
            result_worksheet.write(n, 6, device_area, format1)
            n += 1

result_workbook.close()

time_end = time.process_time()
time_sum = round((time_end - time_start), 2)
counter = int(time_sum / 0.01)

for k in tqdm(range(counter)):
    time.sleep(0.01)

print('数据处理完毕！')

exit_flag = input('按任意键退出程序...')
